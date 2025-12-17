# fund_split_app.py
# -*- coding: utf-8 -*-
from __future__ import annotations

from io import BytesIO
import re
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# -----------------------------
# 설정값 (VBA 로직 그대로)
# -----------------------------
PATTERNS = {
    "연구기금": r"^\(임의_연구\)",
    "장학기금": r"^\(임의_장학\)",
    "건축기금": r"^\(임의_건축\)",
    "특목기금": r"^\(임의_기타\)",
}

DROP_E_VALUES = {
    "미지급금", "미수금", "임의연구기금", "임의건축기금", "예금이자", "예금",
    "임의장학기금", "임의특정목적기금"
}

# VBA에서 지우던 열(열문자 기준). 파이썬에서는 “열문자→인덱스”로 드롭
DROP_COL_LETTERS = ["H", "K", "L", "M", "O", "P", "U", "Y", "Z", "AA"]

# VBA에서 정렬 기준: Q열(=재원). 파이썬은 헤더 "재원" 우선, 없으면 Q fallback
SORT_HEADER = "재원"
SORT_FALLBACK_LETTER = "Q"

# VBA에서 숫자 변환/서식: L, M
NUM_COL_LETTERS = ["L", "M"]


# -----------------------------
# 유틸
# -----------------------------
def col_letter_to_index(letter: str) -> int:
    """A->0, B->1, ... AA->26"""
    letter = letter.upper()
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def safe_strip(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def try_parse_number(x):
    """문자열 '1,234' -> 1234"""
    if x is None:
        return x
    if isinstance(x, (int, float)):
        return x
    s = str(x).strip().replace(",", "")
    if s == "":
        return None
    try:
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s)
    except Exception:
        pass
    return x


def autofit_openpyxl(ws):
    """
    헤더 + 데이터 전체 기준으로 열 너비 자동 조정
    (글자 안 잘리게, 과도하게 넓어지지 않게 제한)
    """
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row, col)
            v = cell.value
            if v is None:
                continue

            # 숫자는 표시 문자열 길이 기준
            if isinstance(v, (int, float)):
                s = f"{v:,.0f}"
            else:
                s = str(v)

            max_len = max(max_len, len(s))

        # 최소/최대 폭 가드
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 70)


def apply_number_format(ws, col_letters):
    for letter in col_letters:
        idx = col_letter_to_index(letter) + 1
        if idx > ws.max_column:
            continue
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, idx)
            cell.value = try_parse_number(cell.value)
            cell.number_format = "#,##0"


# -----------------------------
# 핵심 로직 (VBA 기금재원정리)
# -----------------------------
def split_and_cleanup(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    1) X열(0-based 23) 기준 (임의_연구/장학/건축/기타) 분류
    2) 열 삭제(H,K,L,M,O,P,U,Y,Z,AA)
    3) E열 값 특정 항목 제거
    4) 재원 정렬(헤더 '재원' 우선, 없으면 Q열)
    """
    # X열 존재 확인
    x_idx = col_letter_to_index("X")
    if x_idx >= len(df.columns):
        raise ValueError("원본에 X열이 없습니다. 원장 기본엑셀 형식인지 확인하세요.")

    x_series = df.iloc[:, x_idx].map(safe_strip)

    # 분류
    out = {}
    for sheet_name, pat in PATTERNS.items():
        mask = x_series.str.contains(pat, regex=True, na=False)
        out[sheet_name] = df.loc[mask].copy()

    # 공통 정리 함수
    def cleanup(one: pd.DataFrame) -> pd.DataFrame:
        # E열 특정값 제거
        e_idx = col_letter_to_index("E")
        if e_idx < len(one.columns):
            e_series = one.iloc[:, e_idx].astype(str)
            one = one[~e_series.isin(DROP_E_VALUES)].copy()

        # 열 삭제(열문자 기준 인덱스 드롭)
        drop_idxs = sorted(
            [col_letter_to_index(c) for c in DROP_COL_LETTERS if col_letter_to_index(c) < len(one.columns)],
            reverse=True
        )
        cols = list(one.columns)
        for di in drop_idxs:
            cols.pop(di)
        one = one.loc[:, cols]

        # 재원 기준 정렬
        if SORT_HEADER in one.columns:
            one = one.sort_values(by=SORT_HEADER, ascending=True, kind="mergesort")
        else:
            q_idx = col_letter_to_index(SORT_FALLBACK_LETTER)
            if q_idx < len(one.columns):
                one = one.sort_values(by=one.columns[q_idx], ascending=True, kind="mergesort")

        return one

    for k in list(out.keys()):
        out[k] = cleanup(out[k])

    return out


def build_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        for name, data in sheets.items():
            data.to_excel(writer, sheet_name=name, index=False)
    bio.seek(0)

    wb = load_workbook(bio)

    # ✅ 여기: 모든 시트를 돌면서 숫자서식 + AutoFit
    for ws in wb.worksheets:
        apply_number_format(ws, NUM_COL_LETTERS)
        autofit_openpyxl(ws)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# -----------------------------
# Streamlit 페이지
# -----------------------------
def run():
    # 상단: 뒤로가기
    back_col, _ = st.columns([1, 5])
    with back_col:
        if st.button("← 메인으로"):
            st.session_state["page"] = "main"
            st.rerun()

    st.title("🧩 기금재원정리 (임의기금 분류)")
    st.write("원장 기본엑셀 파일을 업로드하면 4개 시트(연구/장학/건축/특목)로 분류해 새 엑셀을 만들어줍니다.")

    st.markdown("""
        - 지원 형식: XLSX, XLSM  

        1. 회계-장부관리-원장 엑셀자료 메뉴 클릭  
        2. 교비비등록금 회계단위를 조회하여 우클릭 후 *기본엑셀*로 저장(엑셀파일x)  
        3. 파일을 업로드하여 검증작업 진행  

        - 오류 시: 파일명/헤더 행/빈 행 여부를 확인
        """)

    up = st.file_uploader("원본 파일 업로드 (.xlsx/.xlsm)", type=["xlsx", "xlsm"])
    if not up:
        st.stop()

    prog = st.progress(0)
    status = st.empty()

    try:
        status.write("📥 파일 읽는 중...")
        df = pd.read_excel(up, sheet_name=0, dtype=object)
        prog.progress(20)

        status.write("🧠 X열 기준 분류 + 정리(열삭제/행삭제/정렬) 중...")
        sheets = split_and_cleanup(df)
        prog.progress(70)

        status.write("📦 결과 엑셀 생성(서식/AutoFit 포함) 중...")
        out_bytes = build_excel_bytes(sheets)
        prog.progress(95)

        status.write("✅ 완료!")
        prog.progress(100)

        # 화면에는 결과 표를 안 보여주고 요약만
        counts = {k: len(v) for k, v in sheets.items()}
        st.info(f"분류 결과: 연구 {counts['연구기금']:,}건 / 장학 {counts['장학기금']:,}건 / 건축 {counts['건축기금']:,}건 / 특목 {counts['특목기금']:,}건")

        st.download_button(
            "📥 분류 결과 엑셀 다운로드",
            data=out_bytes,
            file_name="기금재원정리_결과.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        prog.progress(100)
        status.write("❌ 오류 발생")
        st.exception(e)
