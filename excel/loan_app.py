# loan_app.py
# -*- coding: utf-8 -*-
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


# ----------------------------- 헤더 -----------------------------
LOAN_HEADERS = [
    "차입금명",
    "회차", "상환예정일", "원금상환예정액", "이자상환예정액", "12.21-31예정액",
    "총이자상환예정액", "예정지연배상금", "예정연체이자", "연체금계산일자", "총 합계",
    "상환일자", "상환원금", "상환이자", "이월이자", "총상환이자",
    "조기상환액", "조기상환이자", "가수금", "지연배상금", "연체금",
    "총 합계", "이자율", "연체유무"
]


# ======================= VBA 매크로 변환 핵심 기능 =======================
def make_loan_workbook(uploaded_files, year_prefix: str) -> BytesIO:

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "통합결과"

    # 헤더 기록
    ws_out.append(LOAN_HEADERS)
    has_data = False

    # 1) 파일들에서 데이터 모으기
    for f in uploaded_files:
        if not f.name.lower().endswith((".xls", ".xlsx", ".xlsm")):
            continue

        file_bytes = BytesIO(f.read())
        wb_src = load_workbook(file_bytes, data_only=True)
        ws_src = wb_src.worksheets[0]

        # 마지막 행(A열 기준)
        last_row = 0
        for r in range(ws_src.max_row, 4, -1):
            if ws_src.cell(row=r, column=1).value not in (None, ""):
                last_row = r
                break

        # 마지막 열(4행 기준)
        last_col = 0
        for c in range(ws_src.max_column, 0, -1):
            if ws_src.cell(row=4, column=c).value not in (None, ""):
                last_col = c
                break

        copy_cols = min(last_col, 23)

        if last_row >= 5 and copy_cols >= 1:
            has_data = True
            base_name = f.name.rsplit(".", 1)[0]

            for r in range(5, last_row + 1):
                row_vals = [base_name]
                empty = True
                for c in range(1, copy_cols + 1):
                    v = ws_src.cell(row=r, column=c).value
                    if v not in (None, ""):
                        empty = False
                    row_vals.append(v)
                if not empty:
                    ws_out.append(row_vals)

    if not has_data:
        output = BytesIO()
        wb_out.save(output)
        output.seek(0)
        return output

    # 2) (옵션) 연도 필터: C열 상환예정일이 year_prefix로 시작하는 것만 남기기
    year_prefix = (year_prefix or "").strip()
    all_rows = [[c for c in row] for row in ws_out.iter_rows(values_only=True)]
    header = all_rows[0]
    body = all_rows[1:]

    if year_prefix:
        body = [
            row for row in body
            if row[2] is not None and str(row[2]).startswith(year_prefix)
        ]

    # 3) 상환예정일로 정렬
    body_sorted = sorted(
        body,
        key=lambda r: str(r[2]) if r[2] is not None else ""
    )

    # 4) 상환예정일별 소계 + 총계 계산
    new_body = []
    current_date = None
    group_acc = None
    total_acc = [0] * len(header)  # 총계용 누적

    def make_subtotal_row(date_value, acc):
        """현재 그룹(acc)에 대한 소계 행 생성"""
        if acc is None:
            return None
        row = [None] * len(header)
        row[0] = "소계"
        row[2] = date_value
        for i in range(3, len(header)):  # D열 이후 숫자 합계
            if acc[i] != 0:
                row[i] = acc[i]
        return row

    for row in body_sorted:
        row_date = row[2]

        # 날짜가 바뀌는 시점에 소계 한 줄 삽입
        if current_date is not None and row_date != current_date:
            subtotal_row = make_subtotal_row(current_date, group_acc)
            if subtotal_row:
                new_body.append(subtotal_row)
            group_acc = None  # 새 그룹 시작

        # 그룹 누적 초기화
        if group_acc is None:
            group_acc = [0] * len(header)

        # 숫자 컬럼 누적 (D열 이후)
        for i, v in enumerate(row):
            if i >= 3 and isinstance(v, (int, float)):
                group_acc[i] += v
                total_acc[i] += v

        new_body.append(list(row))
        current_date = row_date

    # 마지막 그룹 소계
    if current_date is not None:
        subtotal_row = make_subtotal_row(current_date, group_acc)
        if subtotal_row:
            new_body.append(subtotal_row)

    # 총계 행
    total_row = [None] * len(header)
    total_row[0] = "총계"
    for i in range(3, len(header)):
        if total_acc[i] != 0:
            total_row[i] = total_acc[i]
    # 소계들 아래에 한 줄 비우고 총계 추가
    new_body.append([None] * len(header))
    new_body.append(total_row)

    # 5) 시트 갈아엎고 다시 쓰기
    wb_out.remove(ws_out)
    ws_out = wb_out.create_sheet("통합결과", 0)
    ws_out.append(header)
    for row in new_body:
        ws_out.append(row)

    # 6) 숫자 서식 D~V 적용
    for row in ws_out.iter_rows(min_row=2, min_col=4, max_col=22):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # 7) 열 너비 자동 조정 + 틀고정
    for col in ws_out.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_out.column_dimensions[col_letter].width = max_len + 2

    ws_out.freeze_panes = "A2"

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output

    # ======================= 연도 필터(C열) =======================
    year_prefix = (year_prefix or "").strip()
    if year_prefix:
        all_rows = [[c for c in row] for row in ws_out.iter_rows(values_only=True)]
        header = all_rows[0]
        body = all_rows[1:]

        filtered = []
        for row in body:
            val = row[2]
            if val is None:
                continue
            if str(val).startswith(year_prefix):
                filtered.append(row)

        wb_out.remove(ws_out)
        ws_out = wb_out.create_sheet("통합결과", 0)
        ws_out.append(header)
        for row in filtered:
            ws_out.append(row)

    # ======================= 숫자 서식 D~V =======================
    for row in ws_out.iter_rows(min_row=2, min_col=4, max_col=22):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # ======================= 열 너비 자동 조정 =======================
    for col in ws_out.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_out.column_dimensions[col_letter].width = max_len + 2

    ws_out.freeze_panes = "A2"

    output = BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output


# ======================= Streamlit 화면 =======================
def run():

    st.title("💰 사학진흥재단 차입금 정리")

    st.write("여러 차입금 엑셀 파일을 업로드하면 `통합결과` 시트를 만들어서 내려줍니다.")

    files = st.file_uploader(
        "차입금 원본 파일 업로드 (여러 개 가능)",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True
    )

    year = st.text_input("정리할 연도 (예: 2025) — 비워두면 전체 포함", value="")

    if st.button("📊 통합 파일 생성"):
        if not files:
            st.warning("먼저 파일을 업로드하세요.")
        else:
            output = make_loan_workbook(files, year)
            st.success("완료되었습니다!")
            st.download_button(
                label="📥 차입금 통합결과 다운로드",
                data=output,
                file_name=f"차입금_통합_{year or '전체'}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )