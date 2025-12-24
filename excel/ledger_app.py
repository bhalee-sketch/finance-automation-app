# ledger_app.py
# -*- coding: utf-8 -*-
import streamlit as st
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# -------------------------------------------------------
#  원장 통합 함수 (진행률 업데이트 기능 포함)
# -------------------------------------------------------
def merge_ledgers_from_workbooks(files, progress_callback=None) -> BytesIO:
    """
    files: UploadedFile 리스트
    progress_callback: (done, total) → None 형태 함수
    """
    twb = Workbook()
    summary_ws = twb.active
    summary_ws.title = "통합"

    copy_row = 1
    Mcol = 13  # 전표번호(M열)

    total = len(files)
    done = 0

    for f in files:
        done += 1
        if progress_callback is not None:
            progress_callback(done, total)

        # 파일 열기 (Streamlit UploadedFile → BytesIO 변환)
        file_bytes = BytesIO(f.read())
        wb = load_workbook(filename=file_bytes, data_only=True)
        ws = wb.worksheets[0]

        # --- AB열 기준 마지막 행 찾기 ---
        ab_idx = 28  # AB = 28번째 열
        last_row = ws.max_row
        while last_row > 1 and (ws.cell(row=last_row, column=ab_idx).value in (None, "")):
            last_row -= 1
        if last_row <= 1:
            continue

        # --- 헤더 마지막 열 찾기 ---
        last_col = ws.max_column
        while last_col > 1 and (ws.cell(row=1, column=last_col).value in (None, "")):
            last_col -= 1

        # --- 헤더 1회만 복사 ---
        if copy_row == 1:
            for col in range(1, last_col + 1):
                summary_ws.cell(row=1, column=col).value = ws.cell(1, col).value
            copy_row = 2

        # --- 본문 복사 ---
        for r in range(2, last_row + 1):
            for c in range(1, last_col + 1):
                val = ws.cell(row=r, column=c).value
                # 전표번호(M열)은 문자열로 강제
                if c == Mcol and val is not None:
                    val = str(val)
                summary_ws.cell(row=copy_row, column=c).value = val
            copy_row += 1

    # -------------------------------------------------------
    #  서식 정리 (열 너비 자동 조정, 회계 서식 적용)
    # -------------------------------------------------------
    max_row = summary_ws.max_row
    max_col = summary_ws.max_column

    # 열 너비 자동 조정
    for col in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            v = summary_ws.cell(row=row, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        summary_ws.column_dimensions[col_letter].width = max_len + 2

    # U~V (21~22열) 회계 서식
    acc_fmt = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
    for col in (21, 22):
        for row in range(2, max_row + 1):
            cell = summary_ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                cell.number_format = acc_fmt

    # -------------------------------------------------------
    #  메모리에 저장 후 반환
    # -------------------------------------------------------
    output = BytesIO()
    twb.save(output)
    output.seek(0)
    return output

# -------------------------------------------------------
#  Streamlit 실행 화면
# -------------------------------------------------------
def run():
    st.title("📘 회계단위별 원장 통합")

    st.markdown("""
        - 지원 형식: XLSX, XLSM  
                
                    **사용 방법**        
        1. 회계-장부관리-원장 엑셀자료 메뉴 클릭
        2. 각 회계단위를 조회하여 우클릭 후 *엑셀파일*로 저장(기본엑셀x)
        3. 파일을 업로드하여 통합작업 진행
                
        - 오류 시: 파일명/헤더 행/빈 행 여부를 확인
        """)

    files = st.file_uploader(
        "각 회계단위 원장 파일을 업로드하세요.",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
        key="ledger_upload",
    )

    if files and st.button("📂 원장 통합 실행"):
        files_list = list(files)

        # Progress bar 준비
        progress_bar = st.progress(0)
        status_text = st.empty()

        # 진행률 업데이트 함수
        def update_progress(done, total):
            pct = int(done / total * 100)
            progress_bar.progress(pct)
            status_text.text(f"{pct}% 진행 중...  ({done}/{total} 파일 처리 완료)")

        # 실제 통합 실행
        merged_file = merge_ledgers_from_workbooks(files_list, update_progress)

        # 완료 표시
        progress_bar.progress(100)
        status_text.text("✅ 원장 통합 완료!")

        # 다운로드 버튼
        st.download_button(
            label="📥 원장 통합.xlsx 다운로드",
            data=merged_file,
            file_name="원장 통합.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
