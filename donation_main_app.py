# donation_main_app.py
# -*- coding: utf-8 -*-
import streamlit as st

def run():
    # 뒤로가기
    back_col, _ = st.columns([1, 5])
    with back_col:
        if st.button("← 메인으로"):
            # 모드 초기화(선택사항)
            st.session_state.pop("donation_mode", None)
            st.session_state["page"] = "main"
            st.rerun()

    st.title("🎁 출연받은재산 정리")
    st.write("재원을 선택하면 이 페이지에서 바로 작업을 실행합니다.")
    st.markdown("""
        - 지원 형식: XLSX, XLSM  

        1. 회계-세무관리-출연받은재산 사용내역 메뉴 클릭  
        2. 회계단위를 조회하여 우클릭 후 *기본엑셀*로 저장(엑셀파일x)  
        3. 파일을 업로드하여 검증작업 진행
        4. 결과- 부서/기부금 시트별로 정리 

        - 오류 시: 파일명/헤더 행/빈 행 여부를 확인
        """)
    st.markdown("---")

    # 모드 선택 버튼
    if "donation_mode" not in st.session_state:
        st.session_state["donation_mode"] = None

    col1, col2 = st.columns(2)
    with col1:
        if st.button("교비비등록금 재원", use_container_width=True):
            st.session_state["donation_mode"] = "gb"
    with col2:
        if st.button("대학원비등록금 재원", use_container_width=True):
            st.session_state["donation_mode"] = "grad"

    mode = st.session_state["donation_mode"]

    if mode is None:
        st.info("위에서 재원을 선택하면 업로드/다운로드 영역이 나타납니다.")
        return

    st.markdown("---")

    # ===== 교비비등록금 =====
    if mode == "gb":
        st.subheader("✅ 교비비등록금 재원 처리")
        up = st.file_uploader("원본 파일 업로드 (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="up_gb")
        if not up:
            st.stop()

        prog = st.progress(0)
        status = st.empty()

        try:
            status.write("📥 처리 중...")
            prog.progress(20)

            out_bytes = process_gb_like_vba(up)

            prog.progress(95)
            status.write("✅ 완료")
            prog.progress(100)

            st.download_button(
                "📥 결과 엑셀 다운로드",
                data=out_bytes,
                file_name="출연받은재산_교비비등록금.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            prog.progress(100)
            st.exception(e)

    # ===== 대학원비등록금 =====
    elif mode == "grad":
        st.subheader("✅ 대학원비등록금 재원 처리")

        up = st.file_uploader("원본 파일 업로드 (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="up_grad")
        if not up:
            st.stop()

        out_bytes = process_grad_like_vba(up)

        st.download_button(
            "📥 결과 엑셀 다운로드",
            data=out_bytes,
            file_name="출연받은재산_대학원비등록금.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

def process_gb_like_vba(file_like) -> bytes:
    """
    VBA: 교비비출연받은재산정리
    - 적요 공란 제거
    - F/H/I/J 열 삭제
    - 부서별 시트 생성
    - 지정기부금 통합
    - 학생지원팀 → CCF / 지정기부금 이동
    - 합계(L열) + AutoFit
    """

    from io import BytesIO
    import re
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    def autofit_all_columns(ws, min_width=10, max_width=80):
        """
        시트의 모든 열을 대상으로:
        - 헤더 + 데이터 기준 최대 문자열 길이 계산
        - 숫자는 콤마 포함 문자열 기준
        """
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                v = cell.value
                if v is None:
                    continue

                if isinstance(v, (int, float)):
                    s = f"{v:,.0f}"
                else:
                    s = str(v)

                max_len = max(max_len, len(s))

            ws.column_dimensions[col_letter].width = min(
                max(max_len + 2, min_width),
                max_width
            )
    # --------------------------
    # 유틸
    # --------------------------

    def excel_col_to_index(letter: str) -> int:
        n = 0
        for ch in letter.upper():
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n - 1

    def safe_strip(x):
        return "" if x is None else str(x).strip()

    def to_number(x):
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).replace(",", "").strip()
        if s == "":
            return None
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s)
        return None

    def autofit(ws):
        for c in range(1, ws.max_column + 1):
            col_letter = get_column_letter(c)
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = f"{v:,.0f}" if isinstance(v, (int, float)) else str(v)
                max_len = max(max_len, len(s))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 70)

    # --------------------------
    # 1) 원본 읽기
    # --------------------------
    df = pd.read_excel(file_like, sheet_name=0, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]

    # 기본 위치 fallback (VBA 기준)
    col_narr = df.columns[excel_col_to_index("P")]
    col_dept = df.columns[excel_col_to_index("I")]
    col_amt  = df.columns[excel_col_to_index("L")]

    # --------------------------
    # 2) 적요 공란 제거
    # --------------------------
    df = df[df[col_narr].map(safe_strip) != ""].copy()

    # --------------------------
    # 3) F/H/I/J 열 삭제 (역순)
    # --------------------------
    drop_letters = ["J", "I", "H", "F"]
    cols = list(df.columns)
    for lt in drop_letters:
        idx = excel_col_to_index(lt)
        if idx < len(cols):
            cols.pop(idx)
    df = df.loc[:, cols].copy()

    # 삭제 후 다시 위치 보정
    col_narr = df.columns[excel_col_to_index("P")]
    col_dept = df.columns[excel_col_to_index("I")]
    col_amt  = df.columns[excel_col_to_index("L")]

    # --------------------------
    # 4) 금액 숫자화
    # --------------------------
    df[col_amt] = df[col_amt].apply(to_number)

    # --------------------------
    # 5) 부서별 시트 분리
    # --------------------------
    specified_depts = [
        "산학연구지원팀", "비서실", "학생지원팀",
        "대외협력팀", "대학교회", "공간환경시스템공학부"
    ]

    sheets = {}
    for dept in sorted(df[col_dept].dropna().unique()):
        part = df[df[col_dept] == dept].copy()
        sheets[str(dept)] = part

    # --------------------------
    # 6) 지정기부금 통합
    # --------------------------
    지정_rows = []
    for k in list(sheets.keys()):
        if k not in specified_depts:
            지정_rows.append(sheets.pop(k))

    sheets["지정기부금"] = (
        pd.concat(지정_rows, ignore_index=True)
        if 지정_rows else df.iloc[0:0].copy()
    )

    # --------------------------
    # 7) 학생지원팀 → CCF / 지정기부금
    # --------------------------
    if "학생지원팀" in sheets:
        stud = sheets["학생지원팀"].copy()
        narr = stud[col_narr].map(safe_strip)

        mask_ccf = narr.str.contains(r"\(지정\)장학 기부금\(CCF\)", na=False)
        ccf_df = stud[mask_ccf].copy()
        stud = stud[~mask_ccf].copy()

        mask_to_지정 = narr.str.contains(r"\(지정\)기타 지정기부금", na=False) | \
                       narr.str.contains(r"\(지정\)총학생회 기부금", na=False)
        to_지정 = stud[mask_to_지정].copy()
        stud = stud[~mask_to_지정].copy()

        sheets["CCF"] = ccf_df
        if not to_지정.empty:
            sheets["지정기부금"] = pd.concat(
                [sheets["지정기부금"], to_지정],
                ignore_index=True
            )

        sheets["학생지원팀"] = stud

    # --------------------------
    # 8) 시트명 변경
    # --------------------------
    rename_map = {
        "학생지원팀": "교비일반장학",
        "공간환경시스템공학부": "공시학부",
        "산학연구지원팀": "연구소기부",
    }
    for old, new in rename_map.items():
        if old in sheets:
            sheets[new] = sheets.pop(old)

    # --------------------------
    # 9) 엑셀 생성
    # --------------------------
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        for name, data in sheets.items():
            data.to_excel(writer, sheet_name=name[:31], index=False)
    bio.seek(0)

    # --------------------------
    # 10) 합계 + AutoFit
    # --------------------------
    def add_sum_row(ws, k_col_letter="K", l_col_letter="L"):
        k_col = excel_col_to_index(k_col_letter) + 1
        l_col = excel_col_to_index(l_col_letter) + 1

        # 기존 합계행 제거
        for rr in range(ws.max_row, 1, -1):
            v = ws.cell(rr, k_col).value
            if isinstance(v, str) and v.strip() == "합계":
                ws.delete_rows(rr, 1)

        last = ws.max_row
        if last >= 2:
            sum_row = last + 1
            ws.cell(sum_row, k_col).value = "합계"
            ws.cell(sum_row, l_col).value = f"=SUM({l_col_letter}2:{l_col_letter}{last})"
            ws.cell(sum_row, l_col).number_format = "#,##0"
            ws.cell(sum_row, k_col).font = ws.cell(sum_row, k_col).font.copy(bold=True)
            ws.cell(sum_row, l_col).font = ws.cell(sum_row, l_col).font.copy(bold=True)

    wb = load_workbook(bio)
    for ws in wb.worksheets:
        # 숫자 서식 (L열)
        l_col = excel_col_to_index("L") + 1
        if l_col <= ws.max_column:
            for rr in range(2, ws.max_row + 1):
                cell = ws.cell(rr, l_col)
                if isinstance(cell.value, str):
                    vv = cell.value.replace(",", "").strip()
                    if re.fullmatch(r"-?\d+(\.\d+)?", vv or ""):
                        cell.value = float(vv)
                cell.number_format = "#,##0"

        # 합계
        add_sum_row(ws, "K", "L")

        # 🔥 모든 열 AutoFit (반드시 루프 안!)
        autofit_all_columns(ws)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

def process_grad_like_vba(file_like) -> bytes:
    """
    VBA: 대학원비출연받은재산정리_Turbo() 파이썬 변환
    - P열 공란 삭제
    - F/H/I/J 열 삭제
    - L열 숫자화 + #,##0
    - 부서별 시트 생성 (헤더에 '부서' 포함 열, 없으면 I열 폴백)
    - '국제법률대학원' 제외 모든 시트 -> '대학원기부금'으로 합치고 삭제
    - 모든 시트 AutoFit(전체 열) + L합계(각 시트 1줄)
    """

    from io import BytesIO
    import re
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # ---------- util ----------
    def excel_col_to_index(letter: str) -> int:
        n = 0
        for ch in letter.upper():
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n - 1

    def safe_strip(x) -> str:
        return "" if x is None else str(x).strip()

    def to_number(x):
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).replace(",", "").strip()
        if s == "":
            return None
        if re.fullmatch(r"-?\d+(\.\d+)?", s):
            return float(s)
        return None

    def find_dept_col_by_header_or_fallback(df: pd.DataFrame) -> str:
        # VBA: 헤더에 '부서'가 포함된 첫 열, 없으면 I열 폴백
        for c in df.columns:
            if "부서" in str(c).strip():
                return c
        # fallback: I열(원코드 가정)
        idx = excel_col_to_index("I")
        if idx >= len(df.columns):
            raise ValueError("부서 열을 찾지 못했고 I열 폴백도 불가능합니다(컬럼 수 부족).")
        return df.columns[idx]

    def add_sum_row(ws, k_col_letter="K", l_col_letter="L"):
        k_col = excel_col_to_index(k_col_letter) + 1
        l_col = excel_col_to_index(l_col_letter) + 1

        # 기존 합계 제거
        for rr in range(ws.max_row, 1, -1):
            v = ws.cell(rr, k_col).value
            if isinstance(v, str) and v.strip() == "합계":
                ws.delete_rows(rr, 1)

        last = ws.max_row
        if last >= 2:
            sum_row = last + 1
            ws.cell(sum_row, k_col).value = "합계"
            ws.cell(sum_row, l_col).value = f"=SUM({l_col_letter}2:{l_col_letter}{last})"
            ws.cell(sum_row, l_col).number_format = "#,##0"
            ws.cell(sum_row, k_col).font = ws.cell(sum_row, k_col).font.copy(bold=True)
            ws.cell(sum_row, l_col).font = ws.cell(sum_row, l_col).font.copy(bold=True)

    def autofit_all_columns(ws, min_width=10, max_width=80):
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                if isinstance(v, (int, float)):
                    s = f"{v:,.0f}"
                else:
                    s = str(v)
                max_len = max(max_len, len(s))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    # ---------- 1) read ----------
    df = pd.read_excel(file_like, sheet_name=0, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]

    # ---------- 2) P열 공란 삭제 ----------
    p_idx = excel_col_to_index("P")
    if p_idx >= len(df.columns):
        raise ValueError("P열 폴백이 불가능합니다(컬럼 수 부족). 원본 파일 형식 확인 필요.")
    col_p = df.columns[p_idx]
    df = df[df[col_p].map(safe_strip) != ""].copy()

    # ---------- 3) F/H/I/J 삭제 (VBA: J,I,H,F 순서) ----------
    drop_letters = ["J", "I", "H", "F"]
    cols = list(df.columns)
    for lt in drop_letters:
        idx = excel_col_to_index(lt)
        if idx < len(cols):
            cols.pop(idx)
    df = df.loc[:, cols].copy()

    # ---------- 4) L열 숫자화 + 서식 ----------
    l_idx = excel_col_to_index("L")
    if l_idx >= len(df.columns):
        raise ValueError("L열 폴백이 불가능합니다(컬럼 수 부족). 원본 파일 형식 확인 필요.")
    col_l = df.columns[l_idx]
    df[col_l] = df[col_l].apply(to_number)

    # ---------- 5) 부서별 분리 ----------
    col_dept = find_dept_col_by_header_or_fallback(df)

    sheets = {}
    # 부서명 공란 제외
    dept_vals = df[col_dept].map(safe_strip)
    for dept in sorted([d for d in dept_vals.unique() if d]):
        sheets[dept] = df[dept_vals == dept].copy()

    # ---------- 7) 대학원기부금 시트 생성 + 합치기 ----------
    # VBA: 국제법률대학원만 남기고 나머지는 대학원기부금에 합친 후 삭제
    donation_name = "대학원기부금"
    keep_name = "국제법률대학원"

    donation_rows = []
    for name in list(sheets.keys()):
        if name != keep_name:
            donation_rows.append(sheets.pop(name))
    sheets[donation_name] = pd.concat(donation_rows, ignore_index=True) if donation_rows else df.iloc[0:0].copy()

    # 국제법률대학원 시트는 존재하면 유지, 없으면 그냥 donation만 남음
    # (VBA와 동일하게 “없어도 에러 내지 않음”)

    # ---------- 8/9/10) write excel ----------
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # 시트 순서: 국제법률대학원 -> 대학원기부금
        if keep_name in sheets:
            sheets[keep_name].to_excel(writer, sheet_name=keep_name[:31], index=False)
        sheets[donation_name].to_excel(writer, sheet_name=donation_name[:31], index=False)
    bio.seek(0)

    wb = load_workbook(bio)

    # L열 NumberFormat + 합계 + AutoFit(전체열)
    for ws in wb.worksheets:
        # L열 숫자 서식
        l_col_num = excel_col_to_index("L") + 1
        if l_col_num <= ws.max_column:
            for rr in range(2, ws.max_row + 1):
                cell = ws.cell(rr, l_col_num)
                if isinstance(cell.value, str):
                    vv = cell.value.replace(",", "").strip()
                    if re.fullmatch(r"-?\d+(\.\d+)?", vv or ""):
                        cell.value = float(vv)
                cell.number_format = "#,##0"

        # 합계(K/L)
        add_sum_row(ws, "K", "L")

        # AutoFit(전체 열)
        ws.cell(1, 1).alignment = ws.cell(1, 1).alignment.copy(vertical="center")
        ws.sheet_view.showGridLines = True
        autofit_all_columns(ws)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
