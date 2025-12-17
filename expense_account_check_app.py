# expense_account_check_app.py
# -*- coding: utf-8 -*-

from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ======================================================
# V열 기준 시트 구성
# ======================================================
GRAD_SHEETS = [
    "대학원비등록금운영비",
    "국제법률대학원",
    "대학원기부금",
    "대학원임의기금",
    "아동양육",
    "최고경영자",
]

GRAD_V_MAP = {
    "대학원비등록금운영비": ["대학원비등록금운영비(하나17804)"],
    "국제법률대학원": [
        "국제법률대여장학금(하나56104)",
        "국제법률장학금(하나55404)",
        "국제법률기타수익(하나57704)",
    ],
    "대학원기부금": ["대학원기부금(하나58304)"],
    "대학원임의기금": ["대학원임의기금지급(하나45704)"],
    "아동양육": ["아동양육상담 부모콜센터_보탬e(농협7628-91)"],
    "최고경영자": ["최고경영자(하나59004)"],
}

KYOBI_SHEETS = [
    "비등록금운영비",
    "지정기부금",
    "임의기금지급",
    "대학교회",
    "기부부동산",
    "교비일반장학",
    "연구소기부금",
    "제네시스랩",
    "그외",
]

KYOBI_V_MAP = {
    "비등록금운영비": ["비등록금운영비(하나20104)"],
    "지정기부금": ["지정기부금(하나32104)"],
    "임의기금지급": ["임의기금지급(하나50204)", "임의기금지급_감가상각(하나69104)"],
    "대학교회": ["대학교회한국어(하나41404)"],
    "기부부동산": ["기부부동산임대(하나59204)"],
    "연구소기부금": ["연구소기부금(하나41104)"],
    "제네시스랩": ["제네시스랩수입(하나57804)"],
    "교비일반장학": [],
}


# ======================================================
# 유틸
# ======================================================
def _safe_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce"
    )


def _excel_col_to_idx(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


# ======================================================
# 공통 후처리 (속도 최적화 핵심)
# ======================================================
def _postprocess_workbook(
    wb,
    delete_letters,
    widths,
    status_text=None,
    progress=None,
    do_filter=True,
):
    delete_idxs = sorted([_excel_col_to_idx(l) + 1 for l in delete_letters], reverse=True)
    total_ws = len(wb.worksheets)

    for wi, ws in enumerate(wb.worksheets, start=1):
        if status_text:
            status_text.text(f"📐 결과 서식 적용 중... ({wi}/{total_ws})")
        if progress:
            progress.progress(85 + int((wi / total_ws) * 14))

        # 1) 열 삭제
        for idx in delete_idxs:
            if idx <= ws.max_column:
                ws.delete_cols(idx)

        # 2) 미지급금 + 차변0 필터
        if do_filter and ws.max_row >= 2:
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            header_txt = [("" if v is None else str(v).strip()) for v in header]

            if "차변" not in header_txt:
                continue

            debit_col = header_txt.index("차변") + 1
            kept_rows = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                e_val = "" if row[4] is None else str(row[4]).strip()
                if e_val != "미지급금":
                    continue
                try:
                    if float(row[debit_col - 1]) != 0:
                        continue
                except Exception:
                    continue
                kept_rows.append(row)

            ws.delete_rows(2, ws.max_row)
            for r in kept_rows:
                ws.append(r)

        # 3) 열 너비
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 4) 숫자 서식
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        header_txt = [("" if v is None else str(v).strip()) for v in header]

        for target in ["차변", "대변"]:
            if target in header_txt:
                col = header_txt.index(target) + 1
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"


# ======================================================
# 대학원 처리
# ======================================================
def build_grad_excel_by_v(uploaded_file, progress, status_text):
    df = pd.read_excel(uploaded_file, sheet_name=0, dtype=object)
    v_series = df.iloc[:, 21].astype(str).str.strip()

    out_buf = BytesIO()
    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        for s in GRAD_SHEETS:
            mask = v_series.isin(GRAD_V_MAP.get(s, []))
            sub = df.loc[mask].copy()

            for c in ["차변", "대변"]:
                if c in sub.columns:
                    sub.loc[:, c] = _safe_numeric(sub[c])
            sub.to_excel(writer, sheet_name=s, index=False)

    out_buf.seek(0)
    wb = load_workbook(out_buf)

    _postprocess_workbook(
        wb,
        delete_letters=["AA", "Z", "Y", "U", "P", "O", "M", "L", "K", "H", "G", "F"],
        widths=[5.75,14.5,8.63,12.38,9.5,10.13,17,8.63,14,10.75,10.75,17,30,33,27.3],
        status_text=status_text,
        progress=progress,
    )

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    progress.progress(100)
    return final


# ======================================================
# 교비 처리
# ======================================================
def build_kyobi_excel_by_v(uploaded_file, progress, status_text):
    df = pd.read_excel(uploaded_file, sheet_name=0, dtype=object)
    v_series = df.iloc[:, 21].astype(str).str.strip()

    out_buf = BytesIO()
    used = pd.Series(False, index=df.index)

    with pd.ExcelWriter(out_buf, engine="openpyxl") as writer:
        for s in KYOBI_SHEETS:
            if s == "그외":
                continue
            targets = KYOBI_V_MAP.get(s, [])
            mask = v_series.isin(targets)
            used |= mask
            sub = df.loc[mask].copy()
            for c in ["차변", "대변"]:
                if c in sub.columns:
                    sub.loc[:, c] = _safe_numeric(sub[c])
            sub.to_excel(writer, sheet_name=s, index=False)

        other = df.loc[~used].copy()
        for c in ["차변", "대변"]:
            if c in other.columns:
                other.loc[:, c] = _safe_numeric(other[c])
        other.to_excel(writer, sheet_name="그외", index=False)

    out_buf.seek(0)
    wb = load_workbook(out_buf)

    _postprocess_workbook(
        wb,
        delete_letters=["AA", "Z", "Y", "U", "P", "O", "M", "L", "K", "H", "G", "F"],
        widths=[5.75,14.5,8.63,12.38,9.5,10.13,17,8.63,10.5,10.75,10.75,23,30,33,22],
        status_text=status_text,
        progress=progress,
    )

    final = BytesIO()
    wb.save(final)
    final.seek(0)
    progress.progress(100)
    return final


# ======================================================
# UI
# ======================================================
def run():
    # 뒤로가기
    back_col, _ = st.columns([1, 5])
    with back_col:
        if st.button("← 메인으로"):
            # 모드 초기화(선택사항)
            st.session_state.pop("donation_mode", None)
            st.session_state["page"] = "main"
            st.rerun()

    st.title("🧾 지출계좌 재원 검증")

    st.markdown("""
        - 지원 형식: XLSX, XLSM  

        1. 회계-장부관리-원장 엑셀자료 메뉴 클릭
        2. 회계단위를 조회하여 우클릭 후 *기본엑셀*로 저장(엑셀파일x)  
        3. 파일을 업로드하여 검증작업 진행
        4. 결과- 지출계좌 시트별로 정리 

        - 오류 시: 파일명/헤더 행/빈 행 여부를 확인
        """)

    mode = st.radio("회계단위 선택", ["교비비등록금", "대학원비등록금"])

    up = st.file_uploader("원본 파일 업로드", type=["xlsx", "xlsm"])
    if not up:
        return

    progress = st.progress(0)
    status = st.empty()

    if mode == "교비비등록금":
        result = build_kyobi_excel_by_v(up, progress, status)
        name = "지출계좌_검증결과_교비.xlsx"
    else:
        result = build_grad_excel_by_v(up, progress, status)
        name = "지출계좌_검증결과_대학원.xlsx"

    st.download_button("📥 결과 다운로드", result, file_name=name)
