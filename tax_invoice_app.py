# tax_invoice_app.py
# -*- coding: utf-8 -*-

import os
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter


# =========================== 공통 유틸 ===========================

def normalize_key(series: pd.Series) -> pd.Series:
    """사업자번호 비교를 위해 숫자만 남기는 정규화"""
    return series.astype(str).str.replace(r"[^0-9]", "", regex=True).str.strip()


def detect_key_index(header_row, candidates):
    """헤더에서 특정 문자열 포함 컬럼 index 찾기"""
    for idx, val in enumerate(header_row):
        txt = str(val)
        for c in candidates:
            if c in txt:
                return idx
    return None


def find_col(df: pd.DataFrame, keywords):
    """해당 키워드를 가진 컬럼의 엑셀 index(1부터)를 찾기"""
    for col in df.columns:
        s = str(col)
        for kw in keywords:
            if kw in s:
                return df.columns.get_loc(col) + 1
    return None


def pick_col_name(df: pd.DataFrame, keywords):
    """키워드 포함 컬럼명을 반환"""
    for col in df.columns:
        s = str(col)
        for kw in keywords:
            if kw in s:
                return col
    return None


def sanitize_headers(header_row):
    """NaN, 중복 컬럼명 정리"""
    new_headers = []
    used = {}
    for i, h in enumerate(header_row):
        if pd.isna(h) or str(h).strip() == "":
            base = f"Unnamed_{i+1}"
        else:
            base = str(h)

        if base in used:
            used[base] += 1
            name = f"{base}_{used[base]}"
        else:
            used[base] = 1
            name = base

        new_headers.append(name)

    return new_headers


def align_columns(ref_df, target_df):
    """매입시트 컬럼 순서를 매출과 동일하게 정렬"""
    if ref_df.empty or target_df.empty:
        return target_df

    ref_cols = list(ref_df.columns)
    for c in ref_cols:
        if c not in target_df.columns:
            target_df[c] = pd.NA

    return target_df[ref_cols]


# =========================== 매입 전용 정리 ===========================

def clean_buy_df(df):
    """
    매입 전용 정리:
    - 공급받는자등록번호 제거
    - 공급자등록번호(B열)로 재배치
    """
    if df.empty:
        return df

    cols = list(df.columns)

    # 공급받는자등록번호 제거
    remove_cols = []
    for c in cols:
        s = str(c)
        if "공급받는자등록번호" in s or ("공급받는자" in s and "등록번호" in s):
            remove_cols.append(c)

    for c in remove_cols:
        if c in cols:
            cols.remove(c)

    # 공급자등록번호 → B열 배치
    supplier = None
    for c in cols:
        if "공급자등록번호" in str(c):
            supplier = c
            break

    if supplier:
        cols.remove(supplier)
        cols.insert(1, supplier)

    return df[cols]


# =========================== 공통 정리 ===========================

def clean_common_df(df):
    """
    매입/매출 공통 정리:
    - Unnamed 삭제
    - 업태/종목 삭제
    - 사업자번호, 거래처명, 발생금액 삭제
    - 매수_y, 공급가액_y, 부가세액 삭제
    - _dup 컬럼 삭제
    """
    if df.empty:
        return df

    drop_cols = []
    for c in df.columns:
        s = str(c)
        if s.startswith("Unnamed_"):
            drop_cols.append(c)
        elif "업태" in s or "종목" in s:
            drop_cols.append(c)
        # 🔽 여기 줄만 추가했다고 보면 됨
        elif s in ["사업자번호", "거래처명", "발생금액",
                   "매수_y", "공급가액_y", "부가세액"]:
            drop_cols.append(c)
        elif s.endswith("_dup"):
            drop_cols.append(c)

    if drop_cols:
        df = df.drop(columns=drop_cols)

    return df


# =========================== 학사 거래처명 재배치 ===========================

def reorder_haksa_vendor(df):
    """사업자번호_학사 바로 오른쪽에 거래처명_학사 배치"""
    if df.empty:
        return df

    cols = list(df.columns)

    if "사업자번호_학사" in cols and "거래처명_학사" in cols:
        cols.remove("거래처명_학사")
        idx = cols.index("사업자번호_학사")
        cols.insert(idx + 1, "거래처명_학사")
        df = df[cols]

    return df


# =========================== 파일 읽기 ===========================

def import_by_pattern(uploaded_files, pattern, start_row_first):
    processed = 0
    skipped = 0
    first = True
    df_list = []

    for f in uploaded_files:
        if pattern in f.name:
            ext = os.path.splitext(f.name)[1].lower()
            if ext == ".xls":
                st.warning(f"{f.name}은 XLS라서 제외됩니다.")
                skipped += 1
                continue

            try:
                f.seek(0)
                raw = pd.read_excel(
                    f,
                    header=None,
                    engine="openpyxl",
                    dtype=str,
                    na_filter=False,        # ✅ 빈칸/마스킹을 NaN으로 덜 바꿈
                    keep_default_na=False
                )

            except Exception as e:
                st.warning(f"{f.name} 읽기 오류: {e}")
                skipped += 1
                continue

            if len(raw) < start_row_first:
                skipped += 1
                continue

            start = start_row_first - 1 if first else start_row_first
            if len(raw) <= start:
                skipped += 1
                continue

            df_list.append(raw.iloc[start:].copy())
            processed += 1
            first = False

    df = pd.concat(df_list, ignore_index=True) if df_list else pd.DataFrame()
    return df, f"{pattern} → 처리 {processed}건 / 건너뜀 {skipped}건"


# =========================== 매칭 로직 ===========================

def connect_by_id(home_df, haksa_df):
    if home_df.empty:
        return pd.DataFrame()

    # 홈택스 헤더/본문
    home_header = sanitize_headers(list(home_df.iloc[0]))
    home_body = home_df.iloc[1:].reset_index(drop=True)
    home_body.columns = home_header

    # 학사 헤더/본문
    if not haksa_df.empty:
        haksa_header = sanitize_headers(list(haksa_df.iloc[0]))
        haksa_body = haksa_df.iloc[1:].reset_index(drop=True)
        haksa_body.columns = haksa_header
    else:
        haksa_body = pd.DataFrame()

    # 홈택스 키 표준화
    key_idx = detect_key_index(home_header, ["공급자등록번호", "사업자등록번호"])
    if key_idx is None:
        key_idx = 1

    key_col = home_body.columns[key_idx]
    if "공급자등록번호" not in str(key_col):
        home_body["공급자등록번호"] = home_body[key_col]
    else:
        if key_col != "공급자등록번호":
            home_body["공급자등록번호"] = home_body[key_col]

    # 홈택스 금액 표준화
    sup = pick_col_name(home_body, ["공급가액"])
    if sup:
        home_body["공급가액"] = home_body[sup]

    tax = pick_col_name(home_body, ["세액"])
    if tax:
        home_body["세액"] = home_body[tax]

    tot = pick_col_name(home_body, ["합계금액", "발생금액"])
    if tot:
        home_body["합계금액"] = home_body[tot]

    # 학사 표준화
    if not haksa_body.empty:
        key_h = pick_col_name(haksa_body, ["사업자번호"])
        haksa_body["사업자번호_학사"] = haksa_body[key_h]

        sup_h = pick_col_name(haksa_body, ["공급가액"])
        if sup_h:
            haksa_body["공급가액_학사"] = haksa_body[sup_h]

        tax_h = pick_col_name(haksa_body, ["세액"])
        if tax_h:
            haksa_body["세액_학사"] = haksa_body[tax_h]

        tot_h = pick_col_name(haksa_body, ["합계금액", "발생금액"])
        if tot_h:
            haksa_body["합계금액_학사"] = haksa_body[tot_h]

        # 학사 거래처명
        vendor_h = pick_col_name(haksa_body, ["거래처명", "상호"])
        if vendor_h:
            haksa_body["거래처명_학사"] = haksa_body[vendor_h]

        # 머지
        home_body["__KEY"] = normalize_key(home_body["공급자등록번호"])
        haksa_body["__KEY"] = normalize_key(haksa_body["사업자번호_학사"])

        # ✅ merge는 한 번만 (indicator 포함)
        merged = pd.merge(
            home_body,
            haksa_body,
            on="__KEY",
            how="left",
            indicator=True
        )

        # ✅ 홈택스에 없는 학사 키만 추출
        home_keys = set(home_body["__KEY"].dropna().astype(str))
        haksa_only = haksa_body[
            haksa_body["__KEY"].notna() &
            ~haksa_body["__KEY"].astype(str).isin(home_keys)
        ].copy()

        if not haksa_only.empty:
            # merged 구조에 맞게 컬럼 보정
            for c in merged.columns:
                if c not in haksa_only.columns and c != "_merge":
                    haksa_only[c] = pd.NA

            # 컬럼 순서 맞추기
            haksa_only = haksa_only[[c for c in merged.columns if c != "_merge"]]
            haksa_only["_merge"] = "haksa_only"

            merged = pd.concat([merged, haksa_only], ignore_index=True)

        # 마무리 정리
        merged = merged.drop(columns=["__KEY", "_merge"])
    
    else:
        merged = home_body.copy()

    return merged


# =========================== 엑셀 수식 ===========================

def display_len(cell) -> int:
    v = cell.value
    if v is None:
        return 0

    if isinstance(v, bool):
        return 4 if v else 5  # TRUE / FALSE

    if isinstance(v, (int, float)):
        fmt = cell.number_format or ""
        if "," in fmt:
            try:
                return len(f"{v:,.0f}")
            except Exception:
                return len(str(v))
        return len(str(v))

    return len(str(v))

def apply_formulas_and_autofit(writer, sheet, df, is_tax=True):
    ws = writer.book[sheet]
    start_row = 2

    col_B = find_col(df, ["공급자등록번호"])
    col_E = find_col(df, ["공급가액"])
    col_F = find_col(df, ["세액"])
    col_G = find_col(df, ["합계금액"])
    col_K = find_col(df, ["사업자번호_학사"])
    col_P = find_col(df, ["공급가액_학사"])
    col_Q = find_col(df, ["세액_학사"])
    col_R = find_col(df, ["합계금액_학사"])

    col_W = df.shape[1] + 1
    col_X = col_W + 1
    col_Y = col_W + 2
    col_Z = col_W + 3

    ws.cell(1, col_W).value = "사업자번호일치"
    if is_tax:
        ws.cell(1, col_X).value = "공급가액차이"
        ws.cell(1, col_Y).value = "세액차이"
        ws.cell(1, col_Z).value = "합계금액차이"
    else:
        ws.cell(1, col_X).value = "공급가액차이"

    last = start_row + len(df) - 1

    # ── 행별로 수식 채우기 ─────────────────────
    for row in range(start_row, last+1):
        if col_B and col_K:
            ws.cell(row, col_W).value = (
                f"=EXACT({get_column_letter(col_B)}{row},"
                f"{get_column_letter(col_K)}{row})"
            )

        if is_tax:
            if col_E and col_P:
                ws.cell(row, col_X).value = (
                    f"={get_column_letter(col_E)}{row}-"
                    f"{get_column_letter(col_P)}{row}"
                )
            if col_F and col_Q:
                ws.cell(row, col_Y).value = (
                    f"={get_column_letter(col_F)}{row}-"
                    f"{get_column_letter(col_Q)}{row}"
                )
            if col_G and col_R:
                ws.cell(row, col_Z).value = (
                    f"={get_column_letter(col_G)}{row}-"
                    f"{get_column_letter(col_R)}{row}"
                )
        else:
            if col_E and col_R:
                ws.cell(row, col_X).value = (
                    f"={get_column_letter(col_E)}{row}-"
                    f"{get_column_letter(col_R)}{row}"
                )

    # ── 숫자 서식: 천 단위 콤마 "#,##0" 적용 ─────────────────
    amount_cols = {
        col for col in [
            col_E, col_F, col_G,   # 홈택스 금액
            col_P, col_Q, col_R,   # 학사 금액
            col_X, col_Y, col_Z    # 차이 계산 열
        ] if col
    }

    for col in amount_cols:
        for row in range(start_row, last+1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = "#,##0"

    # ── 열 너비 자동 맞춤 (표시값 기준) ─────────────────
    max_col = col_Z if is_tax else col_X

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row in range(1, last+1):  # 헤더 포함
            cell = ws.cell(row=row, column=col_idx)
            max_len = max(max_len, display_len(cell))

        if max_len > 0:
            ws.column_dimensions[col_letter].width = max_len + 2    

    # ── 🔒 공급가액차이 열 고정 폭 (105px ≈ width 15) ─────────
    SUPPLY_DIFF_WIDTH = 15  # 105px 정도
    ws.column_dimensions[get_column_letter(col_X)].width = SUPPLY_DIFF_WIDTH            

def apply_to_all_sheets(writer, sheet_df_map, tax_sheets):
    """
    sheet_df_map: {시트명: df}
    tax_sheets: 세금계산서 시트명 set
    """
    for sheet_name, df in sheet_df_map.items():
        is_tax = sheet_name in tax_sheets
        apply_formulas_and_autofit(
            writer=writer,
            sheet=sheet_name,
            df=df,
            is_tax=is_tax
        )

# =========================== Streamlit UI (run 함수) ===========================

def run():
    """메인 앱(app.py)에서 불러오는 세금계산서 대조 페이지"""
    st.title("🧾 학사시스템과 홈택스 세금계산서 대조")

    uploaded_files = st.file_uploader(
        "세금계산서 관련 8개 파일을 업로드하세요. ex)학사매입세금계산서, 홈택스매출계산서",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
    )
    if not uploaded_files:
        st.info("파일을 업로드하면 매칭 결과가 표시됩니다.")
        return

    patterns = [
        ("홈택스매입세금계산서", 9),
        ("학사매입세금계산서", 1),
        ("홈택스매출세금계산서", 9),
        ("학사매출세금계산서", 1),
        ("홈택스매입계산서", 9),
        ("학사매입계산서", 1),
        ("홈택스매출계산서", 9),
        ("학사매출계산서", 1),
    ]

    data_map = {}
    #st.subheader("파일 로딩 결과")
    for pat, sr in patterns:
         df, msg = import_by_pattern(uploaded_files, pat, sr)
         data_map[pat] = df
         #st.write(msg)

    # 매칭
    buy_tax = connect_by_id(
        data_map["홈택스매입세금계산서"], data_map["학사매입세금계산서"]
    )
    sell_tax = connect_by_id(
        data_map["홈택스매출세금계산서"], data_map["학사매출세금계산서"]
    )
    buy_bill = connect_by_id(
        data_map["홈택스매입계산서"], data_map["학사매입계산서"]
    )
    sell_bill = connect_by_id(
        data_map["홈택스매출계산서"], data_map["학사매출계산서"]
    )

    # 매입 → 매출 구조 맞추기
    buy_tax = align_columns(sell_tax, buy_tax)
    buy_bill = align_columns(sell_bill, buy_bill)

    # 매입 전용 정리 + 공통 정리
    buy_tax = clean_buy_df(buy_tax)
    buy_tax = clean_common_df(buy_tax)
    buy_tax = reorder_haksa_vendor(buy_tax)

    buy_bill = clean_buy_df(buy_bill)
    buy_bill = clean_common_df(buy_bill)
    buy_bill = reorder_haksa_vendor(buy_bill)

    # 매출 공통 정리
    sell_tax = clean_common_df(sell_tax)
    sell_tax = reorder_haksa_vendor(sell_tax)

    sell_bill = clean_common_df(sell_bill)
    sell_bill = reorder_haksa_vendor(sell_bill)

    # 미리보기
    # st.subheader("미리보기")
    # col1, col2 = st.columns(2)
    # with col1:
    #     st.caption("매입세금계산서")
    #     st.dataframe(buy_tax.head())
    #     st.caption("매입계산서")
    #     st.dataframe(buy_bill.head())
    # with col2:
    #     st.caption("매출세금계산서")
    #     st.dataframe(sell_tax.head())
    #     st.caption("매출계산서")
    #     st.dataframe(sell_bill.head())

    # 엑셀 다운로드
    st.subheader("통합 엑셀 다운로드")
    if st.button("📥 대조결과 엑셀 생성"):
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_map = {
                "매입세금계산서_매칭": (buy_tax, True),
                "매출세금계산서_매칭": (sell_tax, True),
                "매입계산서_매칭":     (buy_bill, False),
                "매출계산서_매칭":     (sell_bill, False),
            }

            for sheet_name, (df, is_tax) in sheet_map.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                apply_formulas_and_autofit(
                    writer=writer,
                    sheet=sheet_name,
                    df=df,
                    is_tax=is_tax
                )

        output.seek(0)
        st.download_button(
            "📗 대조결과 파일 다운로드",
            output,
            file_name="세금계산서_통합.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
